import glob
import os
import shutil
import time
from datetime import datetime

import pandas as pd
import schedule

from openpyxl import load_workbook
from openpyxl.packaging import workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Protection
#from openpyxl.shared import SharedWorkbook
from openpyxl import Workbook
import schedule
import time
from datetime import datetime, timedelta
import warnings




# Ignorer les avertissements liés à openpyxl
warnings.simplefilter("ignore", UserWarning)

pd.set_option('display.max_rows', 1000)
pd.set_option('display.max_columns', 1000)
pd.set_option('display.expand_frame_repr', False)


def action():
    # Chemin du dossier où se trouvent les classeurs Excel
    # dossier_source = r'C:\Users\Adam Bamba\Desktop\Projet Power BI\Dossier test impayés'
    # dossier_source2 = r'C:\Users\Adam Bamba\Desktop\Projet Power BI\Dossier test États Locatifs'

    dossier_source = r"G:\Drive partagés\12. Property Management BtoB\8. Reportings\Impayés"
    dossier_source2 = r"G:\Drive partagés\12. Property Management BtoB\8. Reportings\États Locatifs"

    # Chemin du dossier où vous souhaitez copier le dernier classeur Excel
    dossier_destination = r'C:\Users\Adam Bamba\Desktop\Projet en cours\Suivi hebdo des impayés'
    dossier_destination2 = r'C:\Users\Adam Bamba\Desktop\Projet en cours\Etat locatifs'

    #### Impayés#################################################################################
    # Liste tous les fichiers Excel dans le dossier source
    liste_fichiers_excel = glob.glob(os.path.join(dossier_source, '*.xlsx'))

    # Vérifie si la liste des fichiers n'est pas vide
    if liste_fichiers_excel:
        # Trie la liste des fichiers par date de modification (le plus récent d'abord)
        liste_fichiers_excel_tries = sorted(liste_fichiers_excel, key=os.path.getmtime, reverse=True)

        # Sélectionne le dernier fichier (le plus récent)
        dernier_fichier_excel = liste_fichiers_excel_tries[0]

        # Construit le chemin complet du fichier source
        chemin_source = os.path.join(dossier_source, dernier_fichier_excel)

        # Obtient le numéro de semaine de la date d'aujourd'hui
        #numero_semaine = datetime.now().strftime('%U')

        date_actuelle = datetime.now() + + timedelta(weeks=1)
        numero_semaine = date_actuelle.strftime('%U')




        # Construit le nouveau nom de fichier avec le numéro de semaine en suffixe
        nom_base, extension = os.path.splitext(os.path.basename(chemin_source))
        nouveau_nom = f'{nom_base}_Semaine{numero_semaine}{extension}'

        # Construit le chemin complet du fichier de destination
        chemin_destination = os.path.join(dossier_destination, nouveau_nom)

        # Vérifie si le fichier existe déjà dans le dossier de destination
        if os.path.exists(chemin_destination):
            print(f'Le fichier {nouveau_nom} existe déjà dans le dossier de destination.')
        else:
            # Copie le fichier vers le dossier de destination
            shutil.copy2(chemin_source, chemin_destination)
            print(f'Le dernier fichier Excel a été copié de {chemin_source} vers {chemin_destination}.')
    else:
        print('Aucun fichier Excel trouvé dans le dossier source.')

        ###Etat locatifs#####################################################################################################

        # Liste tous les fichiers Excel dans le dossier source

    liste_fichiers_excel2 = glob.glob(os.path.join(dossier_source2, '*.xlsx'))

    # Vérifie si la liste des fichiers n'est pas vide
    if liste_fichiers_excel2:
        # Trie la liste des fichiers par date de modification (le plus récent d'abord)
        liste_fichiers_excel_tries2 = sorted(liste_fichiers_excel2, key=os.path.getmtime, reverse=True)

        # Sélectionne le dernier fichier (le plus récent)
        dernier_fichier_excel2 = liste_fichiers_excel_tries2[0]

        # Construit le chemin complet du fichier source
        chemin_source2 = os.path.join(dossier_source2, dernier_fichier_excel2)

        # Obtient le numéro de semaine de la date d'aujourd'hui
        #numero_semaine2 = datetime.now().strftime('%U')

        date_actuelle = datetime.now() + + timedelta(weeks=1)
        numero_semaine2 = date_actuelle.strftime('%U')


        # Construit le nouveau nom de fichier avec le numéro de semaine en suffixe
        nom_base2, extension2 = os.path.splitext(os.path.basename(chemin_source2))
        nouveau_nom2 = f'{nom_base2}_Semaine{numero_semaine2}{extension2}'

        # Construit le chemin complet du fichier de destination
        chemin_destination2 = os.path.join(dossier_destination2, nouveau_nom2)

        # Vérifie si le fichier existe déjà dans le dossier de destination
        if os.path.exists(chemin_destination2):
            print(f'Le fichier {nouveau_nom2} existe déjà dans le dossier de destination.')
        else:
            # Copie le fichier vers le dossier de destination
            shutil.copy2(chemin_source2, chemin_destination2)
            print(f'Le dernier fichier Excel a été copié de {chemin_source2} vers {chemin_destination2}.')
    else:
        print('Aucun fichier Excel trouvé dans le dossier source.')

    ########## Chargement et retraitemnt des colonnes

    df = pd.read_excel(chemin_destination)

    df_EL = pd.read_excel(chemin_destination2)

    df["Date Compta"] = pd.to_datetime(df["Date Compta"])

    for i in df.index:
        if pd.notna(df.loc[i, "Date Compta"]):
            df.loc[i, "duree"] = (datetime.now() - df.loc[i, "Date Compta"]).days
            df.loc[i, "Semaine"] = datetime.now().isocalendar()[1]
            df.loc[i, "Semaine"] = df.loc[i, "Semaine"].astype(int)
        else:
            df.loc[i, "duree"] = 0
            df.loc[i, "Semaine"] = datetime.now().isocalendar()[1]

    df["duree"] = df["duree"].astype(int)
    df["Semaine"] = df["Semaine"].astype(int)

    for i in df.index:
        if pd.notna(df.loc[i, "Date Compta"]):
            jours_ecoules = (datetime.now() - df.loc[i, "Date Compta"]).days
            if jours_ecoules <= 30:
                df.loc[i, "Cadencement"] = 30
            elif 30 < jours_ecoules <= 90:
                df.loc[i, "Cadencement"] = 90
            elif 90 < jours_ecoules <= 180:
                df.loc[i, "Cadencement"] = 180
            else:
                df.loc[i, "Cadencement"] = "Antérieur"
        else:
            df.loc[i, "Cadencement"] = "Antérieur"

    df["Code mandat"] = ""
    for i in df.index:
        # Convertir la valeur en chaîne de caractères
        locataire_str = str(df.loc[i, "N° Locataire"])

        if len(locataire_str) >= 4:
            df.loc[i, "Code mandat"] = locataire_str[:4]

    for i in df.index:
        if (df.loc[i, "Libellé"] == "Solde créditeur") and (df.loc[i, "Montant encaissé (€)"] < 0) and (
                df.loc[i, "Montant quittancé (€)"] == -df.loc[i, "Montant encaissé (€)"]):
            df.loc[i, "Solde  (€)"] = df.loc[i, "Montant encaissé (€)"].round(2)
            df.loc[i, "Montant quittancé (€)"] = 0
        else:
            df.loc[i, "Solde  (€)"] = df.loc[i, "Montant quittancé (€)"].round(2) - df.loc[
                i, "Montant encaissé (€)"].round(
                2)

    df["Montant quittancé (€)"] = df["Montant quittancé (€)"].round(2)
    df["Montant encaissé (€)"] = df["Montant encaissé (€)"].round(2)
    df["Solde  (€)"] = df["Solde  (€)"].round(2)

    df1 = df[df["Libellé"] == "Solde créditeur"]
    print(df)
    print(df1)

    # Créer un dictionnaire de correspondance entre "N° Locataire" et "Nom locataire"
    correspondance_locataires = df.dropna(subset=["Nom locataire"]).set_index("N° Locataire")["Nom locataire"].to_dict()

    # Appliquer la correspondance pour remplir les cellules vides
    df["Nom locataire"] = df["N° Locataire"].map(correspondance_locataires).fillna(df["Nom locataire"])

    ########## Identification des nouveaux locatires

    # chemin_fichier_destination = r'C:\Users\Adam Bamba\Desktop\Projet en cours\Suivi Impayés - All mandats - 28112023 test.xlsx'

    chemin_fichier_destination = r'G:\Drive partagés\12. Property Management BtoB\8. Reportings\Suivi Impayés - All mandats.xlsx'

    # Charger le DataFrame depuis le fichier destination
    df_destination = pd.read_excel(chemin_fichier_destination, sheet_name='Suivi Impayés', skiprows=3)

    # Identifier les numéros de locataires non présents dans le fichier destination
    numeros_non_identifies = df[~df['N° Locataire'].isin(df_destination['N° Locataire'])]['N° Locataire'].unique()
    print(numeros_non_identifies)

    # Charger le DataFrame depuis le fichier destination
    df_destination2 = pd.read_excel(chemin_fichier_destination, sheet_name='Mapping locataires')

    # Identifier les numéros de locataires non présents dans le fichier destination
    numeros_non_identifies2 = df[~df['N° Locataire'].isin(df_destination2['N° Locataire'])]['N° Locataire'].unique()
    # nom_identifies2 = df[~df['N° Locataire'].isin(df_destination2['N° Locataire'])]['Nom locataire']

    print(numeros_non_identifies2)
    # print(nom_identifies2)

    ########## Mise à jour des onglets

    # Charger le classeur Excel avec openpyxl
    wb = load_workbook(chemin_fichier_destination)

    # Sélectionner la feuille 'Suivi Impayés'
    sheet = wb['Suivi Impayés']

    # Trouver la dernière ligne vide dans la colonne 'N° Locataire'
    derniere_ligne_vide = sheet.max_row
    for row in reversed(range(5, sheet.max_row + 1)):
        if sheet.cell(row=row, column=5).value is not None:  # Utiliser le numéro de colonne 4 pour 'D'
            derniere_ligne_vide = row  # Ajouter 1 pour éviter l'écart
            break

    # Copier les numéros de locataires non identifiés à la dernière ligne vide
    for i, numero_locataire in enumerate(numeros_non_identifies, start=1):
        sheet.cell(row=derniere_ligne_vide + i, column=5,
                   value=numero_locataire)  # Utiliser le numéro de colonne 4 pour 'D'

        sheet.cell(row=derniere_ligne_vide + i, column=5).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                             fill_type="solid")

    ############################################################################################################################

    # Sélectionner la feuille 'Suivi Impayés'
    sheet = wb['Mapping locataires']

    # Trouver la dernière ligne vide dans la colonne 'N° Locataire'
    derniere_ligne_vide = sheet.max_row
    for row in reversed(range(2, sheet.max_row + 1)):
        if sheet.cell(row=row, column=1).value is not None:  # Utiliser le numéro de colonne 4 pour 'D'
            derniere_ligne_vide = row  # Ajouter 1 pour éviter l'écart
            break

    # Copier les numéros de locataires non identifiés à la dernière ligne vide
    for i, numero_locataire2 in enumerate(numeros_non_identifies2, start=1):
        sheet.cell(row=derniere_ligne_vide + i, column=1,
                   value=numero_locataire2)  # Utiliser le numéro de colonne 4 pour 'D'

        # Obtenir la valeur correcte pour la colonne 2 en fonction du numéro de locataire
        nom_identifies2 = df.loc[df['N° Locataire'] == numero_locataire2, 'Nom locataire'].values[0]

        sheet.cell(row=derniere_ligne_vide + i, column=2, value=nom_identifies2)

        sheet.cell(row=derniere_ligne_vide + i, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                             fill_type="solid")
        sheet.cell(row=derniere_ligne_vide + i, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                             fill_type="solid")

    # Sauvegarder les modifications dans le fichier Excel
    wb.save(chemin_fichier_destination)

    ########## finalisation des étapes

    # Charger le classeur Excel avec openpyxl une seule fois en dehors de la boucle
    wb = load_workbook(chemin_fichier_destination)

    # Sélectionner la feuille 'Suivi Impayés'
    sheet = wb['Suivi Impayés']

    # Grouper par numéro de locataire et numéro de semaine, puis calculer la somme des soldes
    somme_solde_par_locataire_semaine = df.groupby(['N° Locataire', 'Semaine'])['Solde  (€)'].sum().reset_index()

    # Commencer à la ligne 6
    for row in df_destination[f"Semaine {numero_semaine}"][5:]:
        row = ""

    # Itérer sur les lignes du DataFrame
    for index, row in somme_solde_par_locataire_semaine.iterrows():
        # Récupérer le numéro de locataire et la semaine de la ligne actuelle
        numero_locataire = row['N° Locataire']
        numero_semaine = int(row['Semaine'])

        # Trouver la colonne correspondante dans la feuille Excel
        colonne_correspondante = f"Semaine {numero_semaine}"
        colonne_correspondante_locataire = 'N° Locataire'

        # Vérifier si la colonne correspondante existe dans le DataFrame
        if colonne_correspondante in df_destination.columns:
            # Récupérer la position (index) de la colonne
            position_colonne = df_destination.columns.get_loc(colonne_correspondante)

            # Récupérer la lettre de la colonne (A, B, C, ...)
            lettre_colonne = get_column_letter(position_colonne + 1)

        if colonne_correspondante_locataire in df_destination.columns:
            # Récupérer la position (index) de la colonne
            position_colonne_locataire = df_destination.columns.get_loc(colonne_correspondante_locataire)

            # Récupérer la lettre de la colonne (A, B, C, ...)
            lettre_colonne_locataire = get_column_letter(position_colonne_locataire + 1)

            # Trouver la ligne correspondante au numéro de locataire dans la feuille Excel
        for row in sheet.iter_rows(min_row=0, max_row=sheet.max_row, min_col=position_colonne_locataire + 1,
                                   max_col=position_colonne_locataire + 1):
            if row[0].value == numero_locataire:
                ligne_locataire = row[0].row
                break
        else:
            ligne_locataire = None  # Indique que le numéro de locataire n'a pas été trouvé

        # Insérer la somme dans la cellule appropriée
        sheet[lettre_colonne + str(ligne_locataire)] = somme_solde_par_locataire_semaine.at[index, 'Solde  (€)']

    print(f"Colonne correspondante : {colonne_correspondante}")
    print(f"Lettre colonne : {lettre_colonne}")
    print(f"Ligne locataire : {ligne_locataire}")

    print(somme_solde_par_locataire_semaine)
    print(somme_solde_par_locataire_semaine.shape)

    ######### Coller la copie de df dans l'onglet "Résultats de la requête"##############

    # Sélectionner la feuille 'Résultats de la requête'
    sheet_resultats = wb['Résultats de la requête']

    # Supprimer toutes les lignes existantes dans la feuille "Résultats de la requête"
    sheet_resultats.delete_rows(1, sheet_resultats.max_row)

    # Trouver la dernière ligne utilisée dans la feuille Excel
    max_row = 1

    # Convertir le DataFrame en une liste de lignes
    rows = list(dataframe_to_rows(df, index=False, header=True))

    # Écrire chaque ligne dans la feuille Excel
    for row in rows:
        sheet_resultats.append(row)
        max_row += 1

    ######### Coller la copie de df dans l'onglet "Etat locatifs"##############

    # Sélectionner la feuille 'EL'
    sheet_resultats2 = wb['EL']

    # Supprimer toutes les lignes existantes dans la feuille "EL"
    sheet_resultats2.delete_rows(1, sheet_resultats2.max_row)

    # Trouver la dernière ligne utilisée dans la feuille Excel
    max_row = 1

    # Convertir le DataFrame en une liste de lignes
    rows = list(dataframe_to_rows(df_EL, index=False, header=True))

    # Écrire chaque ligne dans la feuille Excel
    for row in rows:
        sheet_resultats2.append(row)
        max_row += 1

    # Sauvegarder les modifications dans le fichier Excel
    wb.save(chemin_fichier_destination)

    # Chemin vers le fichier Excel
    file_path = chemin_fichier_destination

    # Chargez le classeur Excel
    wb = load_workbook(file_path)

    # Activez le mode partagé (hérité) pour le classeur
    #wb.security = SharedWorkbook()

    # Sélectionnez la feuille de calcul
    sheet = wb.active

    # Parcourez les colonnes
    for column in sheet.columns:
        # Vérifiez si le libellé de la colonne est égal à "Commentaire"
        if column[3].value != "Commentaires":
            # Verrouillez toutes les cellules de la colonne sauf la première ligne (en-tête)
            for cell in column[1:]:
                cell.protection = Protection(locked=True)

    # Activez le mode partagé du classeur
    workbook.shared_workbook = True

    # Sauvegardez le classeur
    wb.save(chemin_fichier_destination)

    print(
        "Toutes les cellules ont été verrouillées, sauf celles dans les colonnes 'Commentaire', et le classeur est en mode partagé.")




    print(os.path.abspath(__file__))
    print("Action lancée à", datetime.now())


def planifier_action():
    # Planifier l'action toutes les 5 minutes
    #schedule.every(2).minutes.do(action)

    # Planifier l'action tous les mardis à 11h
    schedule.every().tuesday.at("11:00").do(action)


# Appeler la fonction pour planifier l'action
planifier_action()

while True:
    schedule.run_pending()
    time.sleep(1)
